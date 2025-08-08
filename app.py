
import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from calc import run_model, COMPONENT_COLS, normalize_weights

# Extras
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode
import plotly.express as px

st.set_page_config(page_title="Portfolio Sizer", layout="wide")

# ---------- THEME / CSS (Grey background, dark red text) ----------
st.markdown("""
<style>
/* Core palette */
:root{ --up:#21c55d; --down:#ef4444; --grid:#3a4149; --muted:#a18f8f; --accent:#8B0000; }

/* Backgrounds */
html, body, [data-testid="stAppViewContainer"] { background-color: #1E2227; }
[data-testid="stHeader"] { background-color: #1E2227; }
[data-testid="stSidebar"] { background-color: #2A3036; border-right:1px solid var(--grid); }

/* Typography colors */
body, p, span, div, label, .stText, .stMarkdown, .st-emotion-cache, .small,
h1, h2, h3, h4, h5, h6, [data-testid="stMetricValue"], [data-testid="stMetricLabel"],
[data-testid="stMetricDelta"], .stDownloadButton, .stButton { 
  color: #8B0000 !important;
  font-family: Inconsolata, ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, monospace;
}

/* Inputs */
input, textarea, select { color: #8B0000 !important; background-color: #2A3036 !important; }
.stNumberInput input { color: #8B0000 !important; }

/* Data editor tweaks */
.stDataFrame, .stDataEditor { color: #8B0000 !important; }

/* AG Grid */
.ag-root-wrapper, .ag-theme-streamlit-dark, .ag-theme-balham-dark {
  color: #8B0000 !important;
  background-color: #1E2227 !important;
}
.ag-header, .ag-header-cell, .ag-header-cell-label {
  color: #8B0000 !important;
}
.ag-cell { color: #8B0000 !important; }

/* Tabs */
.stTabs [data-baseweb="tab-list"] button p { color: #8B0000 !important; }

/* Metrics row spacing */
.block-container{padding-top:1.0rem; padding-bottom:0.6rem;}
h1,h2,h3{letter-spacing:.4px; font-weight:600}
.small {font-size:0.85rem; color: #8B0000; opacity: .85;}
.topn-bar { border-left: 3px solid var(--accent) !important; }

/* Download buttons */
.stDownloadButton button { color: #8B0000 !important; background:#2A3036 !important; border:1px solid #3a4149 !important; }
.stButton button { color: #8B0000 !important; background:#2A3036 !important; border:1px solid #3a4149 !important; }
</style>
""", unsafe_allow_html=True)

st.title("Portfolio Sizer — Terminal UI (Grey / Dark Red)")

# ---------- SIDEBAR (Global controls) ----------
with st.sidebar:
    st.header("Load / Start")
    uploaded = st.file_uploader("Upload Excel (optional)", type=["xlsx"])

    st.caption("If you don't upload, you'll start with a minimal template you can edit here.")

    st.header("Portfolio Parameters")
    top_n = st.number_input("Top N positions", min_value=1, max_value=100, value=12, step=1)
    min_pos = st.number_input("Min position (%)", min_value=0.0, max_value=100.0, value=3.0, step=0.25) / 100.0
    max_pos = st.number_input("Max position (%)", min_value=0.0, max_value=100.0, value=10.0, step=0.25) / 100.0
    k_offset = st.number_input("Confidence offset (K2)", value=-40.0, step=1.0)

    st.header("Advanced")
    show_weights = st.checkbox("Edit component weights (AC2..AM2)", value=False)
    normalize_final = st.checkbox("Normalize final weights to 100%", value=False)

# ---------- Defaults / Loaders ----------
def default_assets():
    rows = [
        {"asset_id": 1, "asset_name": "Asset 1"},
        {"asset_id": 2, "asset_name": "Asset 2"},
        {"asset_id": 3, "asset_name": "Asset 3"},
    ]
    df = pd.DataFrame(rows)
    for c in COMPONENT_COLS:
        df[c] = 3.0
    return df

def default_cases(assets):
    recs = []
    for aid in assets["asset_id"]:
        recs.extend([
            {"asset_id": aid, "case_label": "Bear", "prob": 0.2, "irr": -0.05},
            {"asset_id": aid, "case_label": "Base", "prob": 0.6, "irr": 0.12},
            {"asset_id": aid, "case_label": "Bull", "prob": 0.2, "irr": 0.25},
        ])
    return pd.DataFrame(recs)

def try_load_excel(file):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb[" Library"] if " Library" in wb.sheetnames else wb.active
        # Parse blocks of 3 rows starting at row 5 until a large empty gap
        assets, cases = [], []
        row = 5
        aid = 1
        empty_streak = 0
        while row < ws.max_row:
            # Read probabilities/IRRs for 3 rows
            U = [ws.cell(row=row+i, column=21).value for i in range(3)]   # U=21
            V = [ws.cell(row=row+i, column=22).value for i in range(3)]   # V=22
            # If the three rows are all empty for U/V, count empty streak
            if all((u in (None, "") and v in (None, "")) for u, v in zip(U, V)):
                empty_streak += 1
                if empty_streak > 10:   # stop after long gap
                    break
                row += 3
                continue
            empty_streak = 0
            # Asset name in B
            name = ws.cell(row=row, column=2).value
            if not isinstance(name, str) or name.strip() == "":
                name = f"Asset {aid}"
            # Component scores AC..AM (29..39)
            comps = []
            for col in range(29, 40):
                val = ws.cell(row=row, column=col).value
                comps.append(float(val) if isinstance(val, (int, float)) else 3.0)
            comp_map = {COMPONENT_COLS[i]: comps[i] for i in range(len(COMPONENT_COLS))}

            arow = {"asset_id": aid, "asset_name": name}
            arow.update(comp_map)
            assets.append(arow)
            # Cases
            labels = ["Bear","Base","Bull"]
            for i in range(3):
                prob = U[i] if isinstance(U[i], (int,float)) else 0.0
                irr  = V[i] if isinstance(V[i], (int,float)) else 0.0
                cases.append({"asset_id": aid, "case_label": labels[i], "prob": float(prob), "irr": float(irr)})
            aid += 1
            row += 3

        # Global params (if present)
        try:
            top_n_xl = ws.cell(row=94, column=7).value
            if isinstance(top_n_xl, (int,float)) and top_n_xl:
                st.sidebar.info(f"Loaded Top-N from Excel: {int(top_n_xl)}")
        except Exception:
            pass

        try:
            k2 = ws.cell(row=2, column=11).value  # K2
            if isinstance(k2, (int,float)):
                st.sidebar.info(f"Loaded K2 offset from Excel: {k2}")
        except Exception:
            pass

        return pd.DataFrame(assets), pd.DataFrame(cases)
    except Exception as e:
        st.warning(f"Could not parse Excel: {e}")
        return default_assets(), default_cases(default_assets())

# Load data
if uploaded:
    assets_df, cases_df = try_load_excel(uploaded)
else:
    assets_df = default_assets()
    cases_df = default_cases(assets_df)

# ---------- Inputs: Assets & Cases ----------
st.subheader("Inputs — Assets")
assets_edit = st.data_editor(
    assets_df,
    column_order=["asset_id","asset_name"] + COMPONENT_COLS,
    disabled=["asset_id"],
    num_rows="dynamic",
    use_container_width=True
)

def ensure_three_cases(cases, assets):
    out = []
    for _, row in assets.iterrows():
        aid = int(row["asset_id"])
        subset = cases[cases["asset_id"] == aid]
        labels = ["Bear","Base","Bull"]
        needed = {l: {"prob": 0.0, "irr": 0.0} for l in labels}
        for _, r in subset.iterrows():
            needed[r["case_label"]] = {"prob": float(r["prob"]), "irr": float(r["irr"])}
        for l in labels:
            out.append({"asset_id": aid, "case_label": l, "prob": needed[l]["prob"], "irr": needed[l]["irr"]})
    return pd.DataFrame(out)

cases_df = ensure_three_cases(cases_df, assets_edit)

st.subheader("Inputs — Cases (per asset)")
cases_edit = st.data_editor(
    cases_df,
    column_order=["asset_id","case_label","prob","irr"],
    num_rows="dynamic",
    use_container_width=True
)

# Component weights (AC2..AM2)
default_w = {c: (0.1 if c in ["AC_ShortFlags","AF_BizModel","AG_TopDown","AH_Competition","AK_Custom1","AL_Custom2"] else
                 0.15 if c in ["AI_Industry"] else 0.05)
             for c in COMPONENT_COLS}
if show_weights:
    st.subheader("Component Weights (AC2..AM2)")
    w_df = pd.DataFrame([default_w])
    w_edit = st.data_editor(w_df, num_rows=1, use_container_width=True)
    weights = w_edit.iloc[0].to_dict()
else:
    weights = default_w

def normalize_probs(df):
    out = df.copy()
    for aid, grp in df.groupby("asset_id"):
        s = grp["prob"].sum()
        if s and abs(s-1.0) > 1e-6:
            out.loc[grp.index, "prob"] = grp["prob"] / s
    return out

cases_norm = normalize_probs(cases_edit)

# ---------- Run model ----------
res = run_model(
    assets_df=assets_edit,
    cases_df=cases_norm,
    comp_weights=weights,
    top_n=top_n,
    min_pos=min_pos,
    max_pos=max_pos,
    k_offset=k_offset,
    normalize=normalize_final
)

# ---------- Top metrics bar ----------
m1, m2, m3, m4 = st.columns(4)
sum_w = res["Weight_Norm" if normalize_final else "Weight"].sum()
top_mask = res["Weight"] > 0
m1.metric("Positions (Top-N)", value=int(top_n))
m2.metric("Sum of Weights", value=f"{sum_w:.1%}")
m3.metric("Max Weight", value=f"{res['Weight'].max():.1%}")
m4.metric("Min Weight (Top-N)", value=f"{res.loc[top_mask,'Weight'].min():.1%}" if top_mask.any() else "—")

# ---------- Three-pane layout ----------
c1, c2, c3 = st.columns([2.1, 2.4, 1.5])

# LEFT PANE — Watchlist grid (AG Grid)
with c1:
    st.subheader("Watchlist")
    quick = st.text_input("Search", placeholder="Filter assets, e.g. 'meta' or '>0.10'", label_visibility="collapsed")
    df_grid = res[["asset_id","asset_name","J_PW_IRR","K_Confidence","Q_Blend","R_RankBlend","Weight","Weight_Norm"]].copy()
    df_grid["TopN"] = (df_grid["R_RankBlend"] <= top_n)

    gb = GridOptionsBuilder.from_dataframe(df_grid)
    gb.configure_pagination(paginationAutoPageSize=True)
    gb.configure_default_column(filter=True, sortable=True, resizable=True,
                                cellStyle={'fontFamily':'Inconsolata, monospace', 'color': "#8B0000"})
    for col in ["J_PW_IRR","K_Confidence","Q_Blend","Weight","Weight_Norm"]:
        gb.configure_column(col, type=["numericColumn"], cellStyle={'textAlign':'right', 'color': "#8B0000"})
    gb.configure_selection(selection_mode="single", use_checkbox=True)
    grid_options = gb.build()
    if quick:
        grid_options["quickFilterText"] = quick

    grid_response = AgGrid(
        df_grid,
        gridOptions=grid_options,
        theme="streamlit",
        update_mode=GridUpdateMode.SELECTION_CHANGED,
        fit_columns_on_grid_load=True
    )
    sel_rows = grid_response["selected_rows"]
    selected_asset = sel_rows[0]["asset_name"] if sel_rows else None

# CENTER PANE — Charts
with c2:
    st.subheader("Chart & Scenario")
    tabs = st.tabs(["J vs K", "Weights", "Diagnostics"])

    # Determine focus set (selected asset -> highlight)
    plot_df = res.copy()
    if selected_asset:
        plot_df["focus"] = np.where(plot_df["asset_name"] == selected_asset, "Selected", "Other")
    else:
        plot_df["focus"] = "Other"

    with tabs[0]:
        fig = px.scatter(
            plot_df, x="J_PW_IRR", y="K_Confidence",
            size="Weight",
            color="focus",
            hover_name="asset_name",
            template="plotly"
        )
        fig.update_layout(paper_bgcolor="#1E2227", plot_bgcolor="#1E2227",
                          font=dict(family="Inconsolata, monospace", color="#8B0000"))
        fig.update_xaxes(gridcolor="#3a4149", zeroline=False, title="PW-IRR (J)",
                         tickfont=dict(color="#8B0000"), title_font=dict(color="#8B0000"))
        fig.update_yaxes(gridcolor="#3a4149", zeroline=False, title="Confidence (K)",
                         tickfont=dict(color="#8B0000"), title_font=dict(color="#8B0000"))
        st.plotly_chart(fig, use_container_width=True)

    with tabs[1]:
        wcol = "Weight_Norm" if normalize_final else "Weight"
        fig2 = px.bar(plot_df.sort_values(wcol, ascending=False), x="asset_name", y=wcol, template="plotly")
        fig2.update_layout(paper_bgcolor="#1E2227", plot_bgcolor="#1E2227",
                           font=dict(family="Inconsolata, monospace", color="#8B0000"),
                           xaxis_title="", yaxis_title=wcol)
        fig2.update_xaxes(tickangle=-30, gridcolor="#3a4149", tickfont=dict(color="#8B0000"))
        fig2.update_yaxes(gridcolor="#3a4149", tickfont=dict(color="#8B0000"), title_font=dict(color="#8B0000"))
        st.plotly_chart(fig2, use_container_width=True)

    with tabs[2]:
        st.write("Intermediate diagnostics")
        st.dataframe(res[["asset_id","asset_name","M_RankJ","N_RankK","O_InvRankJ","P_InvRankK","Q_Blend","R_RankBlend"]],
                     use_container_width=True)

# RIGHT PANE — Portfolio controls & downloads
with c3:
    st.subheader("Portfolio")
    st.markdown(f"**Top-N**: {int(top_n)}  \n**Min%**: {min_pos:.1%}  \n**Max%**: {max_pos:.1%}  \n**K2**: {k_offset:+.0f}")

    # Download buttons
    def to_excel_bytes(df):
        from pandas import ExcelWriter
        import xlsxwriter
        output = BytesIO()
        with ExcelWriter(output, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False, sheet_name="Portfolio")
        return output.getvalue()

    csv_bytes = res.to_csv(index=False).encode("utf-8")
    xl_bytes = to_excel_bytes(res)

    st.download_button("Download results (CSV)", data=csv_bytes, file_name="portfolio_results.csv", mime="text/csv", use_container_width=True)
    st.download_button("Download results (Excel)", data=xl_bytes, file_name="portfolio_results.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

    st.caption("Tip: Edit tables, tweak Top-N / Min% / Max% / K2 on the left. Results refresh automatically.")
